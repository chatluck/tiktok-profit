"""
TikTok 结算文件转换工具
----------------------
用 openpyxl read_only=False 正确读取 TikTok 导出的结算账单和订单列表 XLSX，
输出清理后的 JSON 文件，供网页工具 settlement-report.html 使用。

用法：
    python convert_settlement.py --income income.xlsx --orders orders.xlsx

输出：
    settlement-data.json  (网页工具直接拖入)
"""

import json, re, os, sys
from pathlib import Path
import openpyxl

GIFT_SKUS = {"1A0009"}


def to_float(v):
    if v in (None, "", "-", "/"):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except:
        return 0.0


def text(v):
    return "" if v is None else str(v).strip()


def parse_sku_list(s):
    items = []
    if not s:
        return items
    for part in str(s).replace("\x08", "").split(";"):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"(.+?)x(\d+)(?:\[.*?\])?$", part)
        if m:
            sku, qty = m.group(1).strip(), int(m.group(2))
        else:
            sku, qty = re.sub(r"\[.*?\]", "", part).strip(), 1
        items.append({"sku": sku, "qty": qty, "isGift": sku in GIFT_SKUS or "赠品" in part})
    return items


def extract_orders(ws):
    headers = [text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    required = ["交易号", "商品SKU列表"]
    for r in required:
        if r not in idx:
            raise ValueError(f"订单列表缺少列: {r}")

    order_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(text(v) == "" for v in row[:5]):
            continue
        tx = text(row[idx["交易号"]])
        if not tx:
            continue
        sku_items = parse_sku_list(row[idx["商品SKU列表"]])
        names_raw = text(row[idx.get("商品名清单", -1)]) if "商品名清单" in idx else ""
        names = names_raw.split(";") if names_raw else []
        display = []
        for n, item in enumerate(sku_items):
            cname = names[n].strip() if n < len(names) else ""
            display.append({"sku": item["sku"], "qty": item["qty"], "name": cname, "isGift": item["isGift"]})
        order_map[tx] = {
            "orderNo": text(row[idx.get("订单编号", -1)]) if "订单编号" in idx else "",
            "status": text(row[idx.get("订单状态", -1)]) if "订单状态" in idx else "",
            "skuRaw": text(row[idx["商品SKU列表"]]),
            "items": display,
        }
    return order_map


def extract_settlement(ws):
    headers = [text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    sidx = {h: i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(text(v) == "" for v in row[:5]):
            continue
        oid = text(row[sidx.get("订单ID/调整单ID", -1)])
        if not oid:
            continue
        r = {}
        for h in headers:
            if h in sidx:
                val = row[sidx[h]] if sidx[h] < len(row) else ""
                r[h] = val
        rows.append(r)
    return rows, headers


def main():
    import argparse
    parser = argparse.ArgumentParser(description="TikTok 结算文件转换工具")
    parser.add_argument("--income", required=True, help="结算账单 XLSX 路径")
    parser.add_argument("--orders", required=True, help="订单列表 XLSX 路径")
    parser.add_argument("--output", default="settlement-data.json", help="输出 JSON 路径")
    args = parser.parse_args()

    income_path = Path(args.income)
    orders_path = Path(args.orders)

    if not income_path.exists():
        print(f"错误：找不到结算文件 {income_path}")
        sys.exit(1)
    if not orders_path.exists():
        print(f"错误：找不到订单列表文件 {orders_path}")
        sys.exit(1)

    print(f"读取结算账单: {income_path.name}")
    # 关键：read_only=False 才能正确读取 TikTok 导出文件
    wb_income = openpyxl.load_workbook(income_path, data_only=True, read_only=False)
    ws_income = wb_income["订单详情"] if "订单详情" in wb_income.sheetnames else wb_income[wb_income.sheetnames[0]]
    settlement_rows, headers = extract_settlement(ws_income)
    print(f"  订单详情: {len(settlement_rows)} 行 ({wb_income.sheetnames[0]})")

    print(f"读取订单列表: {orders_path.name}")
    wb_orders = openpyxl.load_workbook(orders_path, data_only=True, read_only=True)
    ws_orders = wb_orders[wb_orders.sheetnames[0]]
    order_map = extract_orders(ws_orders)
    print(f"  订单: {len(order_map)} 个交易号")

    # 合并
    output = {
        "incomeFile": income_path.name,
        "ordersFile": orders_path.name,
        "headers": headers,
        "settlement": settlement_rows,
        "orderMap": order_map,
    }

    out_path = Path(args.output)
    out_path.write_text(json.dumps(output, ensure_ascii=False, default=str), encoding="utf-8")
    print(f"\n已生成: {out_path} ({out_path.stat().st_size / 1024:.0f} KB)")
    print("\n在网页工具 settlement-report.html 中拖入此 JSON 文件即可使用。")


if __name__ == "__main__":
    main()

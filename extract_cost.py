# 从 0421 成本 Excel 重新提取所有产品成本和名称
# 输出 products-data-0421.json

import json
from pathlib import Path
import openpyxl

SRC = Path(r"E:/控价/0421 - 测试成本使用.xlsx")
OUT = Path(r"E:/项目/利润/products-data-0421.json")

wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

def extract_sheet(ws, market):
    headers = [str(c.value) if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}
    
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(str(v).strip() == "" for v in row[:5]):
            continue
        main_sku = str(row[9]).strip() if row[9] else ""
        sub_sku = str(row[10]).strip() if row[10] else ""
        name = str(row[11]).strip() if row[11] else ""
        spec = str(row[12]).strip() if row[12] else ""
        size = str(row[13]).strip() if row[13] else ""
        
        purchase_cost = float(row[34]) if row[34] not in (None, "") else 0.0
        operate_fee = float(row[35]) if row[35] not in (None, "") else 0.0
        packaging_cost = float(row[37]) if row[37] not in (None, "") else 0.0
        shipping_cost = float(row[38]) if row[38] not in (None, "") else 0.0
        cost_cny = float(row[39]) if row[39] not in (None, "") else 0.0
        cost_local = float(row[40]) if len(row) > 40 and row[40] not in (None, "") else 0.0
        
        # 确定 sku：优先用子SKU
        sku = sub_sku if sub_sku else main_sku
        
        cat1 = str(row[6]).strip() if row[6] else ""
        cat2 = str(row[7]).strip() if row[7] else ""
        cat3 = str(row[8]).strip() if row[8] else ""
        status = str(row[26]).strip() if len(row) > 26 and row[26] else "正常"
        
        products.append({
            "market": market,
            "purchaseCost": purchase_cost,
            "dailyPrice": 0,
            "size": size,
            "id": market + sku,
            "spec": spec,
            "name": name,
            "category3": cat3,
            "costCNY": cost_cny,
            "costLocal": cost_local,
            "category2": cat2,
            "operateFee": operate_fee,
            "category1": cat1,
            "sku": sku,
            "shippingCost": shipping_cost,
            "packagingCost": packaging_cost,
            "salesStatus": status,
        })
    return products

all_products = []

# 泰国
ws_th = wb["泰国成本测算"]
th_products = extract_sheet(ws_th, "泰国")
all_products.extend(th_products)
print(f"泰国产品: {len(th_products)}")

# 菲律宾 - 列布局同泰国（col 40=销售成本人民币, col 41=销售成本国家币）
ws_ph = wb["菲律宾成本测算"]
for row in ws_ph.iter_rows(min_row=2, values_only=True):
    if not row or all(str(v).strip() == "" for v in row[:5]):
        continue
    main_sku = str(row[9]).strip() if row[9] else ""
    sub_sku = str(row[10]).strip() if row[10] else ""
    name = str(row[11]).strip() if row[11] else ""
    spec = str(row[12]).strip() if row[12] else ""
    size = str(row[13]).strip() if row[13] else ""
    
    purchase_cost = float(row[35]) if row[35] not in (None, "") else 0.0
    operate_fee = float(row[36]) if row[36] not in (None, "") else 0.0
    packaging_cost = float(row[38]) if row[38] not in (None, "") else 0.0
    shipping_cost = float(row[39]) if row[39] not in (None, "") else 0.0
    cost_cny = float(row[40]) if row[40] not in (None, "") else 0.0
    cost_local = float(row[41]) if len(row) > 41 and row[41] not in (None, "") else 0.0
    rate = float(row[34]) if row[34] not in (None, "") else 0.0
    
    sku = sub_sku if sub_sku else main_sku
    cat1 = str(row[6]).strip() if row[6] else ""
    cat2 = str(row[7]).strip() if row[7] else ""
    cat3 = str(row[8]).strip() if row[8] else ""
    status = str(row[26]).strip() if len(row) > 26 and row[26] else "正常"
    
    all_products.append({
        "market": "菲律宾",
        "purchaseCost": purchase_cost,
        "dailyPrice": 0,
        "size": size,
        "id": "菲律宾" + sku,
        "spec": spec,
        "name": name,
        "category3": cat3,
        "costCNY": cost_cny,
        "costLocal": cost_local,
        "category2": cat2,
        "operateFee": operate_fee,
        "category1": cat1,
        "sku": sku,
        "shippingCost": shipping_cost,
        "packagingCost": packaging_cost,
        "salesStatus": status,
    })

print(f"菲律宾产品: {len(all_products) - len(th_products)}")
print(f"总产品: {len(all_products)}")

result = {"count": len(all_products), "products": all_products}
OUT.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")
print(f"已写入: {OUT}")
